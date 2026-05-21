#!/usr/bin/env python3
"""
compare_ocb_audit.py
────────────────────
Confluence _2025/_2026 Fortify 진단이력과 palantir OCB 진단계획을 비교하여
Excel 비교 문서를 생성합니다.

테이블 구조 (Confluence 진단이력 페이지):
  행0(헤더1): NO | 진단대상 | | 진단담당자 | | FORTIFY | | SNYK | | 요청 | | 비고
  행1(헤더2): 서비스 명칭 | project_key | repository_key | branch_name | build_target | | | 빌드 | 스캔 | 감사 | 리포트 | | 스캔 | 감사 | 리포트 | | 요청자 | 요청일자 | |
  행2~:       데이터

사용법:
    python3 tools/compare_ocb_audit.py
    python3 tools/compare_ocb_audit.py --raw docs/confluence_audit_raw.json
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from html.parser import HTMLParser

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl 미설치: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PALANTIR_DIR = Path(__file__).parent.parent
PLAN_PATH    = PALANTIR_DIR / "docs" / "ocb_scan_plan.md"

# OCB 관련 프로젝트 키 집합 (ocb_scan_plan.md 기준 — Bitbucket 프로젝트 키)
OCB_PROJECTS = {
    "OCBWEBVIEW", "OCBSUGAR", "OCBRWD", "LIVECM",
    "OCBE", "OEP", "OB", "OSA", "OCBX", "OCBNFT", "OCBPASS", "OCBPU",
    "MDPOCBPVD", "OKICK", "OL", "OCB-GAME", "OCB-THP", "OTH", "OC",
    "OVS", "LEAFLET", "OCB-MINT", "AI", "BRG", "MKTIS", "OCBLU",
    "OCBPP", "OCBSR", "OCBLOCK", "OE", "OI", "THOFR_OCB_SYRUP",
    "OCB_BACK_END", "OCBNXIF",
}

# 2025-12 이후 Confluence 테이블은 프로젝트 키 대신 서비스 표시명을 사용.
# 이 기간의 모든 진단 대상은 OCB 서비스 군에 속함 (사용자 확인).
OCB_FROM_MONTH = "2025-12"

# 프로젝트 표시명 → 정규화된 프로젝트 키 매핑
# (Confluence 2025-12 이후 테이블에서 관찰된 값 기준)
PROJECT_NAME_NORMALIZE: dict[str, str] = {
    "ocb webview":              "OCBWEBVIEW",
    "ocb-nft":                  "OCBNFT",
    "ocb nft":                  "OCBNFT",
    "ocb game":                 "OCB-GAME",
    "ocb 캐시백게임":            "OCB-THP",
    "ocb 캐쉬백게임":            "OCB-THP",
    "ocb 운세":                 "OCB-THP",      # OCB 운세, 스타일업 등 → OCB-THP
    "ocb 후원하기":              "OCBE",         # OCB 후원/이벤트 계열
    "ocb 간편결제":              "OCBPU",        # OZ Pay → OCBPU
    "oz pay":                  "OCBPU",
    "ocb_fun_real":             "OCB-THP",
    "ocb oneid pass":           "OCBPASS",
    "oneidpass":                "OCBPASS",
    "ocb 가트결제 (oz pay)":    "OCBPU",
    "ozpay":                    "OCBPU",
    "ocbpayui-frontend-web":    "OCBPU",
    "ocb push":                 "OCB_BACK_END",
    "ocb 전시관리":              "OSA",
    "ocb 이벤트, 오킥게임 등":   "OCBE",
    "오킥 (okick)":              "OKICK",
    "gws":                      "OCBWEBVIEW",  # 오키클럽(GWS) — OCBWEBVIEW 산하
    "talks":                    "OCBWEBVIEW",  # OCB 커뮤니티 talks
}


def _normalize_project(raw_proj: str, month: str) -> str:
    """
    Confluence 표시명을 Bitbucket 프로젝트 키로 정규화.
    - 이미 알려진 키면 그대로 반환
    - 매핑 테이블로 변환 시도
    - 2025-12 이후이면 'OCB' 접두사 키워드가 있는 경우 원본 유지 (is_ocb 로직에서 처리)
    """
    if not raw_proj:
        return raw_proj
    upper = raw_proj.strip().upper()
    if upper in OCB_PROJECTS:
        return upper
    lower = raw_proj.strip().lower()
    for pattern, canonical in PROJECT_NAME_NORMALIZE.items():
        if pattern in lower:
            return canonical
    # 2025-12 이후 미매핑 → 원본 그대로 (is_ocb는 월 기준으로 True)
    return raw_proj.strip()

CLR = {
    "hdr_dark":   "1F2937",
    "hdr_mid":    "374151",
    "done":       "D1FAE5",
    "partial":    "FEF3C7",
    "none":       "FEE2E2",
    "plan_only":  "EFF6FF",
    "conf_only":  "F5F3FF",
    "both":       "ECFDF5",
    "white":      "FFFFFF",
    "row_even":   "F9FAFB",
    "ocb_hi":     "FFF9C4",
}


# ─────────────────────────────────────────────────────────────────────────────
# HTML 파서 — 셀 텍스트 + 링크 동시 추출
# ─────────────────────────────────────────────────────────────────────────────
class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables: list[list[list]] = []  # [table][row][cell] = (text, hrefs)
        self._d = 0          # table depth
        self._cur_table: list[list] = []
        self._cur_row:   list = []
        self._cur_text:  list[str] = []
        self._cur_hrefs: list[tuple[str, str]] = []  # (project, repo)
        self._in_cell = False
        self._cur_href: str | None = None

    def handle_starttag(self, tag, attrs):
        ad = dict(attrs)
        if tag == "table":
            self._d += 1
            if self._d == 1:
                self._cur_table = []
        elif tag == "tr" and self._d == 1:
            self._cur_row = []
        elif tag in ("td", "th") and self._d == 1:
            self._in_cell = True
            self._cur_text = []
            self._cur_hrefs = []
        elif tag == "a" and "href" in ad:
            self._cur_href = ad["href"]
        elif tag == "br" and self._in_cell:
            self._cur_text.append(" ")

    def handle_endtag(self, tag):
        if tag == "table":
            if self._d == 1 and self._cur_table:
                self.tables.append(self._cur_table)
            self._d -= 1
        elif tag == "tr" and self._d == 1:
            if self._cur_row:
                self._cur_table.append(self._cur_row)
        elif tag in ("td", "th") and self._d == 1:
            text = " ".join("".join(self._cur_text).split())
            self._cur_row.append((text, list(self._cur_hrefs)))
            self._in_cell = False
        elif tag == "a":
            self._cur_href = None

    def handle_data(self, data):
        if self._in_cell:
            self._cur_text.append(data)
            if self._cur_href:
                m = re.search(r"projects/([^/]+)/repos/([^/?#]+)", self._cur_href)
                if m:
                    self._cur_hrefs.append((m.group(1).upper(), m.group(2).lower()))


def _cell_text(cell) -> str:
    return cell[0] if cell else ""


def _cell_repos(cell) -> list[tuple[str, str]]:
    return cell[1] if cell else []


# ─────────────────────────────────────────────────────────────────────────────
# Confluence 진단이력 파싱
# ─────────────────────────────────────────────────────────────────────────────
def _parse_month_label(title: str) -> str:
    """'2025-10 Fortify 진단 내역' → '2025-10'"""
    m = re.search(r"(20\d\d[-./]\d{1,2})", title)
    return m.group(1) if m else title[:7]


def _detect_col_indices(rows: list[list]) -> tuple[dict, int]:
    """
    헤더 행을 분석하여 (컬럼 인덱스 dict, 데이터 시작 행 인덱스)를 반환.

    Format A — 단순 포맷 (2026): 헤더 1행
        '' | Project | Repository | Branch | build_target | 비고
    Format B — 상세 포맷 (2025): 헤더 2행
        row0: NO | 진단대상 | | 담당자 | | FORTIFY | SNYK | 요청 | 비고
        row1: 서비스명칭 | project_key | repository_key | branch_name | ...
    """
    if not rows:
        return {}, 1

    def _find(header: list[str], keywords: list[str]) -> int | None:
        for kw in keywords:
            for i, h in enumerate(header):
                if kw in h:
                    return i
        return None

    h0 = [_cell_text(c).strip().lower() for c in rows[0]]

    # Format A 감지: 헤더0에 "project"와 "repository" 키워드가 있으면 단순 포맷
    if _find(h0, ["project"]) is not None and _find(h0, ["repository"]) is not None:
        col = {
            "no":        0,
            "service":   None,   # 단순 포맷에는 서비스명 컬럼 없음
            "proj":      _find(h0, ["project"]) or 1,
            "repo":      _find(h0, ["repository"]) or 2,
            "branch":    _find(h0, ["branch"]) or 3,
            "date":      _find(h0, ["date", "요청일"]),
            "requester": _find(h0, ["requester", "요청자"]),
            "note":      _find(h0, ["비고", "note", "comment", "fortify"]),
            "auditor":   None,
        }
        return col, 1

    # Format B: 헤더 2행 (sub-header에서 컬럼명 찾기)
    # sub-header(h1)에는 NO 컬럼이 없어서 data 행 기준으로 +1 오프셋 필요.
    # h1[k]="project_key" → 실제 data 행에서는 col k+1.
    if len(rows) < 2:
        return {}, 1
    h1 = [_cell_text(c).strip().lower() for c in rows[1]]

    def _find1(keywords: list[str], default: int | None) -> int | None:
        idx = _find(h1, keywords)
        return (idx + 1) if idx is not None else default

    col = {
        "no":        0,
        "service":   _find1(["서비스", "서비스명", "서비스 명"], 1),
        "proj":      _find1(["project_key", "project"],         2),
        "repo":      _find1(["repository_key", "repo"],         3),
        "branch":    _find1(["branch"],                         4),
        "date":      _find1(["요청일자", "요청 일자", "date", "요청접수"], None),
        "requester": _find1(["요청자"],                          None),
        "note":      _find1(["비고", "note", "comment"],        None),
        "auditor":   _find1(["담당", "auditor"],                None),
    }
    return col, 2


def _fix_encoding(s: str) -> str:
    """
    PowerShell ConvertTo-Json 이 UTF-8 한국어를 Latin-1로 잘못 저장한 경우 복원.
    ex) 'ì§\x84ë\x8b¨ëì' → '진단내역'
    """
    if not s:
        return s
    try:
        return s.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return s


def parse_audit_records(raw_path: Path) -> list[dict]:
    """
    Confluence JSON → 진단이력 레코드 리스트.
    """
    with open(raw_path, encoding="utf-8") as f:
        raw = json.load(f)

    records: list[dict] = []

    def _process(page: dict, parent_month: str = ""):
        title   = _fix_encoding(page.get("title", ""))
        html    = _fix_encoding(page.get("body_view", "") or "")
        page_url = page.get("url", "")
        month   = _parse_month_label(title) or parent_month

        parser = TableParser()
        parser.feed(html)

        for tbl in parser.tables:
            if len(tbl) < 2:
                continue

            col, data_start = _detect_col_indices(tbl)

            for row in tbl[data_start:]:
                if not row:
                    continue

                no_val  = _cell_text(row[col.get("no", 0)]) if len(row) > col.get("no", 0) else ""
                if not no_val or not re.match(r"^\d+$", no_val.strip()):
                    continue  # 소계/합계 행 스킵

                def _g(key: str, default="") -> str:
                    idx = col.get(key)
                    if idx is None or idx >= len(row):
                        return default
                    return _cell_text(row[idx]).strip()

                service = _g("service")
                proj_raw = _g("proj")
                repo    = _g("repo").lower()
                branch  = _g("branch")
                date    = _g("date")
                requester = _g("requester")
                note    = _g("note")
                auditor = _g("auditor")
                # build_target: Format A는 col 4, Format B는 별도 컬럼 없음
                build_target = _cell_text(row[4]).strip() if len(row) > 4 else ""

                # href에서 project/repo 보완 (텍스트 파싱 실패 대비)
                if not proj_raw or not repo:
                    for ci in range(min(6, len(row))):
                        for proj_h, repo_h in _cell_repos(row[ci]):
                            if not proj_raw:
                                proj_raw = proj_h
                            if not repo:
                                repo = repo_h

                if not repo:
                    continue

                # 프로젝트명 정규화 (표시명 → Bitbucket 키)
                proj = _normalize_project(proj_raw, month)

                # OCB 판정: 프로젝트 키 매칭 OR 2025-12 이후 전체
                is_ocb = (proj.upper() in OCB_PROJECTS) or (month >= OCB_FROM_MONTH)

                records.append({
                    "no":           no_val,
                    "service":      service,
                    "project":      proj,
                    "proj_raw":     proj_raw,
                    "repo":         repo,
                    "build_target": build_target,
                    "branch":       branch,
                    "date":         _normalize_date(date),
                    "requester":    requester,
                    "auditor":      auditor,
                    "note":         note,
                    "month":        month,
                    "source_page":  title,
                    "page_url":     page_url,
                    "is_ocb":       is_ocb,
                })

    for page in raw.get("pages", []):
        month_label = _parse_month_label(page.get("title", ""))
        _process(page, month_label)
        for sub in page.get("sub_pages", []):
            _process(sub, month_label)

    # 브랜치명 집합 — repo 컬럼에 잘못 파싱된 경우 필터링
    _BRANCH_NAMES = {
        "master", "main", "develop", "development", "dev",
        "production", "prod", "staging", "stage", "release",
        "hotfix", "feature", "test", "qa", "latest",
    }

    # 노이즈 필터: repo 슬러그 형식 검증 (영문/숫자/하이픈/언더스코어만)
    def _valid_repo(r: dict) -> bool:
        repo = r["repo"]
        # 유효한 repo 슬러그: 영문·숫자·-·_ 만 허용, 최소 2자 이상
        if not re.match(r"^[a-zA-Z0-9][a-zA-Z0-9_\-\.]{1,80}$", repo):
            return False
        # 올바른 월 레이블: YYYY-MM 형식만 유지 (부모 페이지 본문 노이즈 제거)
        if not re.match(r"^\d{4}-\d{2}$", r["month"]):
            return False
        # 브랜치명이 repo 컬럼에 들어간 경우 제거
        if repo.lower() in _BRANCH_NAMES:
            return False
        return True

    records = [r for r in records if _valid_repo(r)]

    # 중복 제거: Fortify 기준 = build_target 단위 1건
    # (project, repo, build_target, month) 조합이 고유 식별자
    # build_target이 없는 경우(프론트엔드 단순 repo 등)는 (project, repo, month)
    seen: set[tuple] = set()
    unique: list[dict] = []
    for r in records:
        bt = r["build_target"] or ""
        key = (r["project"], r["repo"], bt, r["month"])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    return unique


def _normalize_date(raw: str) -> str:
    raw = raw.strip()
    m = re.match(r"(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m2 = re.match(r"^(\d{4})(\d{2})(\d{2})$", raw)
    if m2:
        return f"{m2.group(1)}-{m2.group(2)}-{m2.group(3)}"
    return raw


# ─────────────────────────────────────────────────────────────────────────────
# palantir 진단계획 파싱
# ─────────────────────────────────────────────────────────────────────────────
def parse_palantir_plan(plan_path: Path) -> list[dict]:
    text = plan_path.read_text(encoding="utf-8")
    records: list[dict] = []
    cur_priority = cur_project = ""

    for line in text.splitlines():
        m = re.match(r"^#{1,6}\s+([Pp][0-9]+)[^:]*:\s+(\S+)", line)
        if m:
            cur_priority = m.group(1).upper()
            cur_project  = re.sub(r"\s.*", "", m.group(2)).strip("(")
            continue
        m2 = re.match(r"^#{1,6}\s+([Pp][0-9]+)\s*[—–-]", line)
        if m2:
            cur_priority = m2.group(1).upper()
            continue
        m3 = re.match(r"^\|\s*`?([a-zA-Z0-9_\-\.]+)`?\s*\|(.+)\|$", line)
        if m3 and re.search(r"[❌✅🔄]", m3.group(2)):
            repo = m3.group(1).strip().lower()
            cells = [c.strip() for c in m3.group(2).split("|")]
            records.append({
                "repo":     repo,
                "priority": cur_priority,
                "project":  cur_project,
                "inj":  cells[0] if len(cells) > 0 else "❌",
                "xss":  cells[1] if len(cells) > 1 else "❌",
                "file": cells[2] if len(cells) > 2 else "❌",
                "data": cells[3] if len(cells) > 3 else "❌",
                "sca":  cells[4] if len(cells) > 4 else "❌",
            })
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Excel 공통 스타일 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_cell(cell, bg="1F2937", fg="FFFFFF", bold=True, wrap=True, align="center"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, name="Malgun Gothic", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


def _data_cell(cell, bg="FFFFFF", align="left"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Malgun Gothic", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _skill_bg(val: str) -> str | None:
    if "✅" in val: return CLR["done"]
    if "🔄" in val: return CLR["partial"]
    if "❌" in val: return CLR["none"]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: 교차 비교 (OCB 이력 × palantir 계획)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_cross(ws, audit_ocb: list[dict], plan_repos: list[dict]):
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36

    # 제목
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "OCB 서비스 군 — Fortify 진단이력(_2025/_2026) × palantir 진단계획 교차 비교"
    _hdr_cell(c, CLR["hdr_dark"], wrap=False)

    hdrs = ["월", "서비스명 (Confluence)", "project_key", "레포 (Confluence)",
            "Build Target", "Branch", "요청일", "요청자",
            "palantir 우선순위", "INJ", "XSS", "FILE", "DATA", "SCA"]
    for ci, h in enumerate(hdrs, 1):
        _hdr_cell(ws.cell(2, ci, h), CLR["hdr_dark"])

    # 팔란티어 repo_map: repo → plan dict
    plan_map = {r["repo"]: r for r in plan_repos}

    for ri, rec in enumerate(sorted(audit_ocb, key=lambda x: (x["month"], x["project"], x["repo"])), 3):
        repo = rec["repo"]
        plan = plan_map.get(repo) or plan_map.get(repo.replace("-", "_")) or plan_map.get(repo.replace("_", "-"))
        bg   = CLR["both"] if plan else CLR["conf_only"]
        even = ri % 2 == 0

        vals = [
            rec["month"],
            rec["service"],
            rec["project"],
            repo,
            rec["build_target"],
            rec["branch"],
            rec["date"],
            rec["requester"],
            plan["priority"] if plan else "—",
            plan["inj"]  if plan else "—",
            plan["xss"]  if plan else "—",
            plan["file"] if plan else "—",
            plan["data"] if plan else "—",
            plan["sca"]  if plan else "—",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci, val)
            cell_bg = CLR["row_even"] if even and bg == CLR["both"] else bg
            _data_cell(cell, cell_bg, "center" if ci in (1, 3, 6, 7, 9, 10, 11, 12, 13, 14) else "left")
            if ci >= 10:
                sb = _skill_bg(str(val))
                if sb:
                    cell.fill = PatternFill("solid", fgColor=sb)

    ws.column_dimensions["A"].width  = 9
    ws.column_dimensions["B"].width  = 28
    ws.column_dimensions["C"].width  = 14
    ws.column_dimensions["D"].width  = 28
    ws.column_dimensions["E"].width  = 24
    ws.column_dimensions["F"].width  = 10
    ws.column_dimensions["G"].width  = 12
    ws.column_dimensions["H"].width  = 18
    ws.column_dimensions["I"].width  = 10
    for col in ["J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 7


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Confluence 전체 이력 (OCB 한정)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_audit(ws, audit_ocb: list[dict]):
    ws.freeze_panes = "A2"
    hdrs = ["월", "서비스명", "project_key", "레포", "Build Target", "branch", "요청일", "요청자", "담당자", "비고", "출처 페이지"]
    for ci, h in enumerate(hdrs, 1):
        _hdr_cell(ws.cell(1, ci, h), CLR["hdr_dark"])

    for ri, r in enumerate(sorted(audit_ocb, key=lambda x: (x["month"], x["project"])), 2):
        bg = CLR["row_even"] if ri % 2 == 0 else CLR["white"]
        for ci, val in enumerate([
            r["month"], r["service"], r["project"], r["repo"],
            r["build_target"], r["branch"], r["date"], r["requester"], r["auditor"],
            r["note"], r["source_page"]
        ], 1):
            _data_cell(ws.cell(ri, ci, val), bg,
                       "center" if ci in (1, 3, 6, 7) else "left")

    for ci, w in enumerate([9, 26, 14, 28, 24, 10, 12, 18, 16, 22, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: palantir 진단계획 (OCB 관련)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_plan(ws, plan_repos: list[dict]):
    ws.freeze_panes = "A2"
    hdrs = ["레포 슬러그", "project", "우선순위", "INJ", "XSS", "FILE", "DATA", "SCA"]
    for ci, h in enumerate(hdrs, 1):
        _hdr_cell(ws.cell(1, ci, h), CLR["hdr_dark"])

    for ri, r in enumerate(sorted(plan_repos, key=lambda x: (x["priority"], x["repo"])), 2):
        bg = CLR["row_even"] if ri % 2 == 0 else CLR["white"]
        for ci, val in enumerate([r["repo"], r["project"], r["priority"],
                                   r["inj"], r["xss"], r["file"], r["data"], r["sca"]], 1):
            _data_cell(ws.cell(ri, ci, val), bg,
                       "left" if ci <= 2 else "center")
            if ci >= 4:
                sb = _skill_bg(str(val))
                if sb:
                    ws.cell(ri, ci).fill = PatternFill("solid", fgColor=sb)

    for ci, w in enumerate([30, 14, 8, 7, 7, 7, 7, 7], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Gap 분석
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_gap(ws, audit_ocb: list[dict], plan_repos: list[dict]):
    ws.freeze_panes = "A3"

    plan_repos_lower = {r["repo"]: r for r in plan_repos}
    audit_repos = {r["repo"] for r in audit_ocb}

    # A) Fortify 이력 있는데 palantir 계획 없음
    gap_a = [r for r in audit_ocb if r["repo"] not in plan_repos_lower]
    # 중복 제거 (repo 기준)
    seen_a: set[str] = set()
    gap_a_uniq = []
    for r in gap_a:
        if r["repo"] not in seen_a:
            seen_a.add(r["repo"])
            gap_a_uniq.append(r)

    # B) palantir P1 계획 있는데 Fortify 이력 없음
    gap_b = [r for r in plan_repos if r["priority"] == "P1" and r["repo"] not in audit_repos]

    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"Gap 분석 — Fortify 이력 有 / palantir 계획 無 및 역방향 누락")
    _hdr_cell(c, CLR["hdr_dark"], wrap=False, align="left")
    row += 1

    # ── 섹션 A ──
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"■ A. Fortify 이력 有 — palantir 계획 未등록  ({len(gap_a_uniq)}건)")
    _hdr_cell(c, CLR["conf_only"], "374151", align="left")
    row += 1
    for ci, h in enumerate(["월", "서비스명", "project_key", "레포", "branch", "요청일", "요청자", "조치권고"], 1):
        _hdr_cell(ws.cell(row, ci, h), CLR["hdr_mid"])
    row += 1
    for r in sorted(gap_a_uniq, key=lambda x: x["project"]):
        for ci, val in enumerate([r["month"], r["service"], r["project"], r["repo"],
                                   r["branch"], r["date"], r["requester"],
                                   "palantir ocb_scan_plan.md 추가 검토 필요"], 1):
            _data_cell(ws.cell(row, ci, val), CLR["conf_only"],
                       "center" if ci in (1, 3, 5, 6) else "left")
        row += 1

    row += 1

    # ── 섹션 B ──
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row, 1, f"■ B. palantir P1 레포 — Fortify 이력 없음  ({len(gap_b)}건)")
    _hdr_cell(c, CLR["plan_only"], "1E3A5F", align="left")
    row += 1
    for ci, h in enumerate(["레포 슬러그", "project", "우선순위", "INJ", "XSS", "FILE", "DATA", "조치권고"], 1):
        _hdr_cell(ws.cell(row, ci, h), CLR["hdr_mid"])
    row += 1
    for r in sorted(gap_b, key=lambda x: x["repo"]):
        for ci, val in enumerate([r["repo"], r["project"], r["priority"],
                                   r["inj"], r["xss"], r["file"], r["data"],
                                   "Fortify 진단이력 신규 등록 필요"], 1):
            _data_cell(ws.cell(row, ci, val), CLR["plan_only"],
                       "left" if ci in (1, 2, 8) else "center")
            if ci in (4, 5, 6, 7):
                sb = _skill_bg(str(val))
                if sb:
                    ws.cell(row, ci).fill = PatternFill("solid", fgColor=sb)
        row += 1

    for ci, w in enumerate([9, 28, 14, 28, 10, 12, 16, 36], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: 전체 Confluence 이력 (OCB + 비OCB)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_all(ws, all_records: list[dict]):
    ws.freeze_panes = "A2"
    hdrs = ["월", "OCB여부", "서비스명", "project_key", "레포", "branch", "요청일", "요청자", "출처"]
    for ci, h in enumerate(hdrs, 1):
        _hdr_cell(ws.cell(1, ci, h), CLR["hdr_dark"])

    for ri, r in enumerate(sorted(all_records, key=lambda x: (x["month"], x["project"])), 2):
        bg = CLR["ocb_hi"] if r["is_ocb"] else (CLR["row_even"] if ri % 2 == 0 else CLR["white"])
        for ci, val in enumerate([
            r["month"],
            "✅ OCB" if r["is_ocb"] else "",
            r["service"], r["project"], r["repo"],
            r["branch"], r["date"], r["requester"], r["source_page"]
        ], 1):
            _data_cell(ws.cell(ri, ci, val), bg,
                       "center" if ci in (1, 2, 4, 6, 7) else "left")

    for ci, w in enumerate([9, 9, 26, 14, 28, 10, 12, 18, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# 마크다운 요약
# ─────────────────────────────────────────────────────────────────────────────
def _build_md(audit_ocb, plan_repos, out_path: Path):
    plan_lower = {r["repo"] for r in plan_repos}
    audit_repos = {r["repo"] for r in audit_ocb}

    gap_a_repos = sorted({r["repo"] for r in audit_ocb if r["repo"] not in plan_lower})
    gap_b = [r for r in plan_repos if r["priority"] == "P1" and r["repo"] not in audit_repos]

    months_seen = sorted({r["month"] for r in audit_ocb})

    lines = [
        "# OCB 보안진단 — Fortify 이력 × palantir 계획 비교 요약",
        f"> 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"> 대상: Confluence _2025/_2026 Fortify 진단이력 × palantir ocb_scan_plan.md",
        "",
        "## 통계",
        f"| 항목 | 수치 |",
        f"|---|---|",
        f"| Confluence OCB 진단이력 (레포 단위) | **{len({r['repo'] for r in audit_ocb})}개** |",
        f"| Confluence OCB 진단이력 (전체 행) | **{len(audit_ocb)}건** |",
        f"| palantir 진단계획 레포 | **{len(plan_repos)}개** |",
        f"| 커버리지 범위 | {', '.join(months_seen)} |",
        f"| Gap A — Fortify 이력 有, 계획 無 | **{len(gap_a_repos)}건** |",
        f"| Gap B — P1 계획 有, Fortify 이력 無 | **{len(gap_b)}건** |",
        "",
        "## A. Fortify 이력 有 — palantir 계획 未등록",
        "",
        "| 레포 | project | 최근 진단월 |",
        "|---|---|---|",
    ]
    repo_month: dict[str, str] = {}
    repo_proj:  dict[str, str] = {}
    for r in audit_ocb:
        if r["repo"] in gap_a_repos:
            if r["repo"] not in repo_month or r["month"] > repo_month[r["repo"]]:
                repo_month[r["repo"]] = r["month"]
                repo_proj[r["repo"]]  = r["project"]
    for repo in gap_a_repos:
        lines.append(f"| `{repo}` | {repo_proj.get(repo,'')} | {repo_month.get(repo,'')} |")

    lines += [
        "",
        "## B. palantir P1 레포 — Fortify 이력 없음",
        "",
        "| 레포 슬러그 | project | INJ | XSS | FILE | DATA | SCA |",
        "|---|---|---|---|---|---|---|",
    ]
    for r in sorted(gap_b, key=lambda x: x["repo"]):
        lines.append(f"| `{r['repo']}` | {r['project']} | {r['inj']} | {r['xss']} | {r['file']} | {r['data']} | {r['sca']} |")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[저장] {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 통계 출력
# ─────────────────────────────────────────────────────────────────────────────
def _print_stats(all_records, audit_ocb, plan_repos):
    print(f"\n{'='*60}")
    print(f" Confluence 전체 이력: {len(all_records)}건 (중복제거)")
    print(f" OCB 해당 레포: {len(audit_ocb)}건 ({len({r['repo'] for r in audit_ocb})}개 레포)")
    print(f" palantir 계획 레포: {len(plan_repos)}개")

    months = sorted({r['month'] for r in all_records})
    print(f" 수집 범위: {months[0]} ~ {months[-1]}")

    proj_cnt: dict[str, int] = {}
    for r in audit_ocb:
        proj_cnt[r["project"]] = proj_cnt.get(r["project"], 0) + 1
    print(f"\n OCB project별 이력 건수 (상위 10):")
    for proj, cnt in sorted(proj_cnt.items(), key=lambda x: -x[1])[:10]:
        print(f"   {proj:20s} {cnt:3d}건")
    print('='*60)


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw",  default="docs/confluence_audit_raw.json")
    ap.add_argument("--plan", default="docs/ocb_scan_plan.md")
    ap.add_argument("--out",  default="")
    args = ap.parse_args()

    raw_path  = PALANTIR_DIR / args.raw
    plan_path = PALANTIR_DIR / args.plan
    date_str  = datetime.now().strftime("%Y%m%d")
    out_xlsx  = PALANTIR_DIR / (args.out or f"docs/ocb_audit_comparison_{date_str}.xlsx")
    out_md    = out_xlsx.with_suffix(".md")

    if not raw_path.exists():
        print(f"[ERROR] {raw_path} 없음 — fetch_confluence_audit_history.ps1 먼저 실행", file=sys.stderr)
        sys.exit(1)

    print(f"[1] Confluence 이력 파싱: {raw_path}")
    all_records = parse_audit_records(raw_path)
    audit_ocb   = [r for r in all_records if r["is_ocb"]]
    print(f"    전체 {len(all_records)}건 → OCB {len(audit_ocb)}건 ({len({r['repo'] for r in audit_ocb})}개 레포)")

    print(f"[2] palantir 계획 파싱: {plan_path}")
    plan_repos = parse_palantir_plan(plan_path)
    print(f"    {len(plan_repos)}개 레포")

    _print_stats(all_records, audit_ocb, plan_repos)

    print(f"\n[3] Excel 생성: {out_xlsx}")
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "교차 비교(OCB)"
    _sheet_cross(ws1, audit_ocb, plan_repos)

    ws2 = wb.create_sheet("Fortify이력_OCB")
    _sheet_audit(ws2, audit_ocb)

    ws3 = wb.create_sheet("palantir계획")
    _sheet_plan(ws3, plan_repos)

    ws4 = wb.create_sheet("Gap분석")
    _sheet_gap(ws4, audit_ocb, plan_repos)

    ws5 = wb.create_sheet("전체이력(OCB+비OCB)")
    _sheet_all(ws5, all_records)

    wb.save(out_xlsx)
    print(f"[저장] {out_xlsx}")

    print(f"[4] 마크다운 요약: {out_md}")
    _build_md(audit_ocb, plan_repos, out_md)

    print(f"\n완료:")
    print(f"  Excel    → {out_xlsx}")
    print(f"  Markdown → {out_md}")


if __name__ == "__main__":
    main()
